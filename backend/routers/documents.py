from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, BackgroundTasks, Form
from fastapi.responses import JSONResponse
import os
import tempfile
from langchain.text_splitter import RecursiveCharacterTextSplitter
from typing import Optional, cast
from pydantic import BaseModel
from pathlib import Path
from datetime import datetime, timedelta
from sentence_transformers import SentenceTransformer
import numpy as np
from backend.models import Document, DocumentChunk, User
from backend.database import get_db
from backend.auth import get_current_user
import uuid

# Initialize embedding model
model_path = os.path.join(os.path.dirname(__file__), '..', 'models', 'all-MiniLM-L6-v2')
embedding_model = SentenceTransformer(model_path)

# Python 3.12-friendly document extractors
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook

import logging
logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/documents",
    tags=["documents"],
    responses={404: {"description": "Not found"}},
)

logger.info("Documents router initialized")

# Document expires after 24 hours by default
DOCUMENT_TTL_HOURS = 24

class DocumentResponse(BaseModel):
    filename: str
    content: str
    content_type: str
    document_id: str
    expires_at: datetime

class DocumentReference(BaseModel):
    conversation_id: str
    document_id: str
    filename: str
    content_type: str

@router.post("")
async def save_document_reference(document_ref: DocumentReference):
    """Save document reference without processing content"""
    db = await get_db()
    
    # Check if document exists
    existing = await db.documents.find_one({"_id": document_ref.document_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Update document with reference info
    await db.documents.update_one(
        {"_id": document_ref.document_id},
        {"$set": {
            "conversation_id": document_ref.conversation_id,
            "filename": document_ref.filename,
            "content_type": document_ref.content_type
        }}
    )
    
    return {"status": "success", "document_id": document_ref.document_id}

@router.post("/upload")
async def upload_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    conversation_id: Optional[str] = Form(None)
):
    """Handle file upload and text extraction"""
    tmp_file_path = None  # Initialize before try block
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    
    filepath = Path(file.filename)
    ext = filepath.suffix.lower()
    if ext not in ['.pdf', '.docx', '.txt', '.xlsx']:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    
    try:
        # Save file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_file:
            content = await file.read()
            if len(content) > 10 * 1024 * 1024:  # 10MB limit
                raise HTTPException(status_code=400, detail="File too large (max 10MB)")
            tmp_file.write(content)
            tmp_file_path = tmp_file.name
        
        # Extract text using Python 3.12-compatible libraries
        def extract_text_from_file(path: str, ext: str) -> str:
            try:
                if ext == '.pdf':
                    text_parts = []
                    reader = PdfReader(path)
                    for page in reader.pages:
                        page_text = page.extract_text() or ''
                        if page_text:
                            text_parts.append(page_text)
                    return '\n'.join(text_parts)
                elif ext == '.docx':
                    doc = DocxDocument(path)
                    paras = [p.text for p in doc.paragraphs if p.text]
                    # Extract text from tables as well
                    for table in getattr(doc, 'tables', []):
                        for row in table.rows:
                            paras.append('\t'.join(cell.text for cell in row.cells))
                    return '\n'.join(paras)
                elif ext == '.xlsx':
                    wb = load_workbook(path, data_only=True)
                    lines = []
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(values_only=True):
                            str_vals = [str(v) for v in row if isinstance(v, (str, int, float)) and v is not None]
                            if str_vals:
                                lines.append('\t'.join(str_vals))
                    return '\n'.join(lines)
                elif ext == '.txt':
                    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                        return f.read()
                else:
                    return ''
            except Exception:
                # Fallthrough on any parsing error
                return ''

        text = extract_text_from_file(tmp_file_path, ext)
        if not text.strip():
            raise HTTPException(status_code=400, detail="Could not extract text from the uploaded file.")
        
        # Chunk text for better handling
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200
        )
        chunks = text_splitter.split_text(text)
        
        # Create document record
        document_id = str(uuid.uuid4())
        expires_at = datetime.utcnow() + timedelta(hours=DOCUMENT_TTL_HOURS)
        db = await get_db()

        # Store document and chunks without embeddings first
        document = {
            "_id": document_id,
            "filename": file.filename,
            "user_email": current_user.email,
            "content": text,
            "content_type": file.content_type or "application/octet-stream",
            "expires_at": expires_at,
            "conversation_id": conversation_id,
            "created_at": datetime.utcnow(),
            "chunk_count": len(chunks)
        }

        # Insert document and then chunks (embeddings to be added in background)
        await db.documents.insert_one(document)
        if chunks:
            await db.document_chunks.insert_many([
                {
                    "document_id": document_id,
                    "chunk_index": i,
                    "content": chunk,
                    "embedding": None,
                    "created_at": datetime.utcnow()
                } for i, chunk in enumerate(chunks)
            ])
        
        # Schedule background task for embedding computation
        if background_tasks:
            background_tasks.add_task(compute_embeddings, db, document_id, chunks)
        else:
            logger.warning("BackgroundTasks not provided, embeddings will not be computed in background")
        
        # Clean up
        os.unlink(tmp_file_path)
        
        return JSONResponse({
            "filename": file.filename,
            "content": text,
            "content_type": file.content_type,
            "document_id": document_id,
            "expires_at": expires_at.isoformat()
        })
        
    except Exception as e:
        # Clean up if error occurs
        if tmp_file_path and os.path.exists(tmp_file_path):
            try:
                os.unlink(tmp_file_path)
            except:
                pass
        raise HTTPException(
            status_code=500,
            detail=f"Error processing file: {str(e)}"
        )

@router.get("/{document_id}", response_model=DocumentResponse)
async def get_document(document_id: str, current_user: User = Depends(get_current_user)):
    """Fetch a document by ID for the current user"""
    db = await get_db()
    doc = await db.documents.find_one({"_id": document_id, "user_email": current_user.email})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return DocumentResponse(
        filename=doc.get("filename", ""),
        content=doc.get("content", ""),
        content_type=doc.get("content_type", "application/octet-stream"),
        document_id=doc.get("_id"),
        expires_at=doc.get("expires_at")
    )

@router.delete("/{document_id}")
async def delete_document(document_id: str, current_user: User = Depends(get_current_user)):
    """Delete a document by ID"""
    db = await get_db()
    # Delete document and its chunks
    result = await db.documents.delete_one({"_id": document_id, "user_email": current_user.email})
    # Only delete chunks if the document delete actually occurred
    if result.deleted_count:
        await db.document_chunks.delete_many({"document_id": document_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"status": "success"}

@router.post("/cleanup")
async def cleanup_documents():
    """Clean up expired documents"""
    db = await get_db()
    now = datetime.utcnow()
    
    # Find expired documents
    expired_docs = await db.documents.find({
        "expires_at": {"$lt": now}
    }).to_list(1000)  # Large enough number to get all expired docs
    
    if not expired_docs:
        return {"status": "success", "deleted": 0}
    
    doc_ids = [doc["_id"] for doc in expired_docs]
    
    # Delete documents and their chunks
    await db.documents.delete_many({"_id": {"$in": doc_ids}})
    await db.document_chunks.delete_many({"document_id": {"$in": doc_ids}})
    
    return {"status": "success", "deleted": len(doc_ids)}

async def compute_embeddings(db, document_id: str, chunks: list):
    """Background task to compute and update chunk embeddings"""
    try:
        # Encode all chunks in a single batch for efficiency, returning numpy array.
        embeddings = cast(np.ndarray, embedding_model.encode(chunks, convert_to_tensor=False))
        # The result is a 2D numpy array, convert to list of 1D lists.
        chunk_embeddings = embeddings.tolist()
        
        # Update chunks with embeddings
        from pymongo import UpdateOne
        operations = [
            UpdateOne(
                {"document_id": document_id, "chunk_index": i},
                {"$set": {"embedding": embedding}}
            ) for i, embedding in enumerate(chunk_embeddings)
        ]
        
        if operations:
            await db.document_chunks.bulk_write(operations)
    except Exception as e:
        logger.error(f"Error in background embedding task: {str(e)}")